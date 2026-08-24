from __future__ import annotations
import argparse, re
from datetime import date, datetime
from pathlib import Path
from typing import Any
from openpyxl import load_workbook
from .db import DB_PATH, connect, initialize

EXPECTED_HEADERS = {
 "RECORD ID":"record_id","BADGE NUMBER":"badge_number","RANK":"rank","LAST NAME":"last_name","FIRST NAME":"first_name","MIDDLE NAME":"middle_name","SUFFIX":"suffix","CAMP":"camp","OFFICE":"office","GENDER":"gender","CLASSIFICATION":"classification","TYPE":"personnel_type","DUPLICATE STATUS":"duplicate_status","DUPLICATE TYPE":"duplicate_type","CREATED AT":"created_at","UPDATED AT":"updated_at","DRIVEFILEID":"drive_file_id","DRIVE FILE ID":"drive_file_id",
 "ID NUMBER":"id_number","BIRTHDATE":"birthdate","CIVIL STATUS":"civil_status","RELIGION":"religion","HIGHEST EDUCATION":"highest_education","HIGHEST ATTAINED EDUCATION":"highest_education","ADDRESS NO":"address_no","ADDRESS STREET":"address_street","ADDRESS BARANGAY":"address_barangay","ADDRESS CITY":"address_city","ADDRESS PROVINCE":"address_province","ADDRESS ZIP":"address_zip","EMERGENCY CONTACT":"emergency_contact","EMERGENCY RELATIONSHIP":"emergency_relationship","EMERGENCY NUMBER":"emergency_number","EMERGENCY ADDRESS":"emergency_address"
}
FIELDS=list(dict.fromkeys(EXPECTED_HEADERS.values()))
INSERT_SQL=f"""INSERT INTO personnel ({', '.join(FIELDS)}, source_order) VALUES ({', '.join(':'+f for f in FIELDS)}, :source_order)
ON CONFLICT(badge_number) DO UPDATE SET {', '.join(f'{f}=excluded.{f}' for f in FIELDS if f!='badge_number')}, source_order=excluded.source_order;"""

def normalize(value:Any)->str|None:
 if value is None:return None
 if isinstance(value,(datetime,date)):return value.isoformat()
 text=str(value).strip();return text or None

def normalize_badge(value:Any)->str|None:
 if value is None:return None
 if isinstance(value,float) and value.is_integer():return str(int(value))
 return normalize(value)

def normalize_drive_file_id(value:Any)->str|None:
 text=normalize(value)
 if not text:return None
 for pattern in [r"/file/d/([A-Za-z0-9_-]+)",r"[?&]id=([A-Za-z0-9_-]+)",r"/d/([A-Za-z0-9_-]+)"]:
  match=re.search(pattern,text)
  if match:return match.group(1)
 return text

def import_workbook(workbook_path:Path|str,db_path:Path|str=DB_PATH,worksheet_name:str="LIST")->int:
 workbook_path=Path(workbook_path)
 if not workbook_path.exists():raise FileNotFoundError(f"Workbook not found: {workbook_path}")
 workbook=load_workbook(workbook_path,read_only=True,data_only=True)
 if worksheet_name not in workbook.sheetnames:raise ValueError(f"Worksheet '{worksheet_name}' was not found.")
 rows=workbook[worksheet_name].iter_rows(values_only=True);raw_headers=next(rows,None)
 if not raw_headers:raise ValueError("Worksheet is empty.")
 headers=[str(v).strip().upper() if v is not None else "" for v in raw_headers];column_map={i:EXPECTED_HEADERS[h] for i,h in enumerate(headers) if h in EXPECTED_HEADERS}
 missing=sorted({"badge_number","last_name","first_name"}-set(column_map.values()))
 if missing:raise ValueError(f"Missing required columns: {', '.join(missing)}")
 initialize(db_path);imported=0
 with connect(db_path) as connection:
  for source_order,values in enumerate(rows,start=1):
   record={field:None for field in FIELDS}
   for index,field in column_map.items():
    if index>=len(values):continue
    value=values[index]
    record[field]=normalize_badge(value) if field=="badge_number" else normalize_drive_file_id(value) if field=="drive_file_id" else normalize(value)
   if not record["badge_number"]:continue
   record["source_order"]=source_order;connection.execute(INSERT_SQL,record);imported+=1
 workbook.close();return imported

def main()->None:
 parser=argparse.ArgumentParser(description="Import NBPattendance personnel into SQLite.");parser.add_argument("workbook");parser.add_argument("--sheet",default="LIST");parser.add_argument("--db",default=str(DB_PATH));args=parser.parse_args();count=import_workbook(args.workbook,args.db,args.sheet);print(f"Imported {count} personnel record(s) into {args.db}")
if __name__=="__main__":main()
